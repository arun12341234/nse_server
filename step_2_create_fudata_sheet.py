#!/usr/bin/env python3
"""
FuData Sheet Creator for ABB Futures
Creates a complete FuData sheet structure with all contracts

Author: Arun - Aryan Tech World Private Limited
Date: December 13, 2025
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from collections import defaultdict

class FuDataSheetCreator:
    """Create FuData sheet with multiple futures contracts"""
    
    def __init__(self, symbol='ABB', lot_size=125):
        self.symbol = symbol
        self.lot_size = lot_size
        self.max_contracts = 10595418
        self.max_contracts_value = 84763.344
        self.start_date = []
        
        # Define all expiry dates for 2025
        self.expiries = [
            # datetime(2025, 1, 30),   # January
            # datetime(2025, 2, 27),   # February
            # datetime(2025, 3, 27),   # March
            # datetime(2025, 4, 24),   # April
            # datetime(2025, 5, 29),   # May
            # datetime(2025, 6, 26),   # June
            # datetime(2025, 7, 31),   # July
            # datetime(2025, 8, 28),   # August
            # datetime(2025, 9, 25),   # September
            # datetime(2025, 10, 30),  # October
            # datetime(2025, 11, 27),  # November
        ]
        self.start_dates_dist = defaultdict(list)
        # Start dates for each contract (roughly 3 months before expiry)
        self.start_dates = [
            # datetime(2024, 11, 1),   # Jan contract starts Nov
            # datetime(2024, 12, 2),   # Feb contract starts Dec
            # datetime(2024, 12, 27),  # Mar contract starts late Dec
            # datetime(2025, 2, 1),    # Apr contract starts Feb
            # datetime(2025, 2, 28),   # May contract starts late Feb
            # datetime(2025, 3, 28),   # Jun contract starts late Mar
            # datetime(2025, 4, 25),   # Jul contract starts late Apr
            # datetime(2025, 5, 30),   # Aug contract starts late May
            # datetime(2025, 6, 27),   # Sep contract starts late Jun
            # datetime(2025, 8, 1),    # Oct contract starts Aug
            # datetime(2025, 8, 29),   # Nov contract starts late Aug
        ]
        
        # Base price and volatility for simulation
        self.base_price = 6500
        self.volatility = 0.02  # 2% daily volatility
        
        self.wb = None
        self.ws = None
        
    def create_workbook(self):
        """Create new Excel workbook"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "FuData"
        print("✓ Created new workbook")
    
    def generate_price_series(self, start_date, end_date, base_price):
        """Generate realistic price series using random walk"""
        dates = pd.date_range(start=start_date, end=end_date, freq='B')  # Business days
        
        prices = [base_price]
        for i in range(1, len(dates)):
            # Random walk with drift
            change = np.random.normal(0, self.volatility) * prices[-1]
            new_price = prices[-1] + change
            # Keep prices in reasonable range
            new_price = max(min(new_price, base_price * 1.3), base_price * 0.7)
            prices.append(new_price)
        
        df = pd.DataFrame({
            'Date': dates,
            'Price': prices
        })
        
        return df
    def generate_contract(self, expiry_date, contract_num):
        # print(expiry_date, contract_num)
        contract_data = []
        for d in self.start_dates_dist[expiry_date]:
            # print(d,"kkk", f"./NSE_Downloads/FO_Bhavcopy/FO_UDiFF_{d.strftime('%Y%m%d')}.csv")
            from pathlib import Path
            file_path = Path(f"./NSE_Downloads/FO_Bhavcopy/FO_UDiFF_{d.strftime('%Y%m%d')}.csv")

            if file_path.exists():
                print(f"{file_path} File exists")
                df = pd.read_csv(file_path)
                # print(df,"@@@@@@@@@@@@@@@")
                df["TradDt"] = pd.to_datetime(df["TradDt"], errors="coerce")
                df["XpryDt"] = pd.to_datetime(df["XpryDt"], errors="coerce")
                # max_expiry = df["XpryDt"].max()
                # print(max_expiry)
                filtered_df = df[
                    (df["TckrSymb"] == f"{self.symbol}") 
                    &
                    (
                        (df["FinInstrmTp"] == "STF") |
                        (df["FinInstrmTp"] == "IDF")
                    )
                ]
                # print(filtered_df)
                # max_expiry = filtered_df["XpryDt"].max()
                # df_max = filtered_df[filtered_df["XpryDt"] == max_expiry]
                # print(
                #     df_max
                # )
                df_max = filtered_df[filtered_df["XpryDt"] == expiry_date]
                row = df_max.iloc[0]
                # print(row)
                # if hasattr(row["TradDt"], "strftime") else row["TradDt"]
                record = {
                    "Instrument": "FUTSTK",
                    "Symbol": row["TckrSymb"],
                    "Date": row["TradDt"].strftime("%d-%b-%Y") ,
                    "Expiry": row["XpryDt"].strftime("%d-%b-%Y"),
                    "Open": float(row["OpnPric"]),
                    "High": float(row["HghPric"]),
                    "Low": float(row["LwPric"]),
                    "Close": float(row["ClsPric"]),
                    "LTP": float(row["LastPric"]),
                    "Settle_Price": float(row["SttlmPric"]),
                    "Contracts": int(row["TtlTradgVol"]),
                    "Turnover_Lacs": int(row["TtlTrfVal"]//100000),
                    "Open_Interest": int(row["OpnIntrst"]),
                    "Change_OI": int(row["ChngInOpnIntrst"]),
                    "OI_Contracts": int(row["TtlTradgVol"]),
                    "Change_OI_Contracts": int(row["TtlTradgVol"]),
                }
                # print(record)
                contract_data.append(record)
                # print(start_date, max_expiry)
                # # if start_date not in self.start_dates:
                # #     self.start_dates.append(start_date)
                # if max_expiry not in self.expiries:
                #     self.expiries.append(max_expiry)
            else:
                pass
                # print(f"{file_path} File does not exist")
            # dates.append(current_date)
            # current_date += timedelta(days=1)
        ##############
        return pd.DataFrame(contract_data)


    def generate_contract_data(self, expiry_date, start_date, contract_num):
        """Generate data for a single futures contract"""
        # print(
        #     self.start_dates,
        #     self.expiries
        # )
        # print(expiry_date, "start_date",start_date,"end_date",self.start_dates[contract_num+2], contract_num)
        # # move to first day of next month safely
        # next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)

        # end_of_month = next_month - timedelta(days=1)
        # # print("end_of_month:", end_of_month)

        # current_date = start_date
        # dates = []
        # contract_data = []

        # while current_date <= end_of_month:
        #     # print(current_date)
        #     # print(
        #     #     f"./NSE_Downloads/FO_Bhavcopy/FO_UDiFF_{current_date.strftime('%Y%m%d')}.csv"
        #     # )
        #     from pathlib import Path

        #     file_path = Path(f"./NSE_Downloads/FO_Bhavcopy/FO_UDiFF_{current_date.strftime('%Y%m%d')}.csv")

        #     if file_path.exists():
        #         print(f"{file_path} File exists")
        #         df = pd.read_csv(file_path)
        #         # print(df)
        #         df["TradDt"] = pd.to_datetime(df["TradDt"], errors="coerce")
        #         df["XpryDt"] = pd.to_datetime(df["XpryDt"], errors="coerce")
        #         # max_expiry = df["XpryDt"].max()
        #         # print(max_expiry)
        #         filtered_df = df[
        #             (df["TckrSymb"] == f"{self.symbol}") 
        #             &
        #             (df["FinInstrmTp"] == "STF")
        #         ]
        #         # print(filtered_df)
        #         max_expiry = filtered_df["XpryDt"].max()
        #         df_max = filtered_df[filtered_df["XpryDt"] == max_expiry]
        #         print(
        #             df_max
        #         )
        #         row = df_max.iloc[0]
        #         # if hasattr(row["TradDt"], "strftime") else row["TradDt"]
        #         record = {
        #             "Instrument": "FUTSTK",
        #             "Symbol": row["TckrSymb"],
        #             "Date": row["TradDt"].strftime("%d-%b-%Y") ,
        #             "Expiry": row["XpryDt"].strftime("%d-%b-%Y"),
        #             "Open": float(row["OpnPric"]),
        #             "High": float(row["HghPric"]),
        #             "Low": float(row["LwPric"]),
        #             "Close": float(row["ClsPric"]),
        #             "LTP": float(row["LastPric"]),
        #             "Settle_Price": float(row["SttlmPric"]),
        #             "Contracts": int(row["TtlTradgVol"]),
        #             "Turnover_Lacs": int(row["TtlTrfVal"]//100000),
        #             "Open_Interest": int(row["OpnIntrst"]),
        #             "Change_OI": int(row["ChngInOpnIntrst"]),
        #             "OI_Contracts": int(row["TtlTradgVol"]),
        #             "Change_OI_Contracts": int(row["TtlTradgVol"]),
        #         }
        #         # print(record)
        #         contract_data.append(record)
        #         # print(start_date, max_expiry)
        #         # # if start_date not in self.start_dates:
        #         # #     self.start_dates.append(start_date)
        #         # if max_expiry not in self.expiries:
        #         #     self.expiries.append(max_expiry)
        #     else:
        #         pass
        #         # print(f"{file_path} File does not exist")
        #     # dates.append(current_date)
        #     current_date += timedelta(days=1)
        # ##############
        # return pd.DataFrame(contract_data)









        # print(
        #     expiry_date.strftime('%Y%m%d')
        #     # , start_date
        # )
        # print(
        #     f"./NSE_Downloads/FO_Bhavcopy/FO_UDiFF_{start_date.strftime('%Y%m%d')}.csv"
        # )
        # for i in range(1,32):
        #     if i < 10:
        #         print(f"0{i}")
        #     else:
        #         print(i)
        
        # # Adjust base price based on time
        # days_from_start = (expiry_date - start_date).days
        # price_adjustment = 1.0 - (contract_num * 0.02)  # Each contract slightly different
        # adjusted_base = self.base_price * price_adjustment
        
        # # Generate price series
        # price_df = self.generate_price_series(start_date, min(expiry_date, datetime.now()), adjusted_base)
        
        # contract_data = []
        
        # for idx, row in price_df.iterrows():
        #     date = row['Date']
        #     settle_price = row['Price']
            
        #     # Generate OHLC around settle price
        #     daily_range = settle_price * 0.015  # 1.5% daily range
        #     open_price = settle_price + np.random.uniform(-daily_range/2, daily_range/2)
        #     high_price = max(open_price, settle_price) + np.random.uniform(0, daily_range/2)
        #     low_price = min(open_price, settle_price) - np.random.uniform(0, daily_range/2)
        #     close_price = settle_price + np.random.uniform(-daily_range/4, daily_range/4)
        #     ltp = close_price
            
        #     # Generate volume and OI
        #     # OI builds up towards expiry, then decreases in last week
        #     days_to_expiry = (expiry_date - date).days
        #     if days_to_expiry > 7:
        #         oi_factor = min(1.0, (len(price_df) - days_to_expiry) / len(price_df))
        #     else:
        #         oi_factor = days_to_expiry / 7.0  # Decrease in last week
            
        #     max_oi = np.random.randint(15000, 25000)
        #     oi_contracts = int(max_oi * oi_factor * np.random.uniform(0.8, 1.2))
        #     oi_contracts = max(oi_contracts, 5)
            
        #     # Calculate OI change
        #     if len(contract_data) > 0:
        #         change_oi_contracts = oi_contracts - contract_data[-1]['OI_Contracts']
        #     else:
        #         change_oi_contracts = oi_contracts
            
        #     # Volume
        #     contracts = np.random.randint(10, 200)
        #     turnover_lacs = (contracts * self.lot_size * settle_price) / 100000
            
        #     # Open Interest in value
        #     open_interest = oi_contracts * self.lot_size * settle_price
        #     change_oi = change_oi_contracts * self.lot_size * settle_price
            
        #     record = {
        #         'Instrument': 'FUTSTK',
        #         'Symbol': self.symbol,
        #         'Date': date,
        #         'Expiry': expiry_date,
        #         'Open': round(open_price, 2),
        #         'High': round(high_price, 2),
        #         'Low': round(low_price, 2),
        #         'Close': round(close_price, 2),
        #         'LTP': round(ltp, 2),
        #         'Settle_Price': round(settle_price, 2),
        #         'Contracts': contracts,
        #         'Turnover_Lacs': round(turnover_lacs, 2),
        #         'Open_Interest': round(open_interest, 2),
        #         'Change_OI': round(change_oi, 2),
        #         'OI_Contracts': oi_contracts,
        #         'Change_OI_Contracts': change_oi_contracts
        #     }
            
        #     contract_data.append(record)
        
        # return pd.DataFrame(contract_data)
    
    def write_contract_header(self, col_start, expiry_date, start_date, contract_df):
        """Write header rows for a contract"""
        
        # Row 1 - Symbol, Lot Size, Max Contracts
        self.ws.cell(row=1, column=col_start, value="Symbol =")
        self.ws.cell(row=1, column=col_start+1, value=self.symbol)
        self.ws.cell(row=1, column=col_start+2, value="Lot Size =")
        self.ws.cell(row=1, column=col_start+3, value=self.lot_size)
        self.ws.cell(row=1, column=col_start+4, value="Max. Contracts =")
        self.ws.cell(row=1, column=col_start+6, value=self.max_contracts)
        self.ws.cell(row=1, column=col_start+7, value=self.max_contracts_value)
        self.ws.cell(row=1, column=col_start+9, value="Opening OI Contracts =")
        self.ws.cell(row=1, column=col_start+14, value="Opening Settle Price")
        
        # Row 2 - Expiry, 90% Max OI, Present OI
        self.ws.cell(row=2, column=col_start+1, value="Expiry =")
        self.ws.cell(row=2, column=col_start+2, value=expiry_date.strftime("%d-%b-%Y"))
        self.ws.cell(row=2, column=col_start+4, value="90% of Max. OI Contracts =")
        self.ws.cell(row=2, column=col_start+7, value=round(self.max_contracts_value * 0.9, 4))
        self.ws.cell(row=2, column=col_start+9, value="Present OI Contracts =")
        self.ws.cell(row=2, column=col_start+12, value=contract_df['OI_Contracts'].iloc[-1])
        self.ws.cell(row=2, column=col_start+13, value=contract_df['Settle_Price'].iloc[-1])
        self.ws.cell(row=2, column=col_start+14, value="Present settle Price")
        
        # Row 3 - Start date, Max OI month
        self.ws.cell(row=3, column=col_start+1, value="Start dt. =")
        if start_date:
            start_date = pd.to_datetime(start_date).strftime("%d-%b-%Y")
        self.ws.cell(row=3, column=col_start+2, value=start_date)
        self.ws.cell(row=3, column=col_start+4, value="Max. OI Contracts month =")
        self.ws.cell(row=3, column=col_start+7, value=contract_df['OI_Contracts'].max())
        self.ws.cell(row=3, column=col_start+9, value="OI Low")
        self.ws.cell(row=3, column=col_start+11, value="Rows =")
        self.ws.cell(row=3, column=col_start+12, value=len(contract_df))
        
        # Row 5 - Column headers
        headers = ['Instrument', 'Symbol', 'Date', 'Expiry', 'Open', 'High', 'Low', 
                  'Close', 'LTP', 'Settle Price', 'No. of contracts', 'Turnover in lacs',
                  'Open Int', 'Change in OI', 'OI Contracts', 'Change in OI Contracts']
        
        for i, header in enumerate(headers):
            cell = self.ws.cell(row=5, column=col_start+i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    def write_contract_data(self, col_start, contract_df):
        """Write data rows for a contract"""
        
        start_row = 6
        
        for idx, row in contract_df.iterrows():
            self.ws.cell(row=start_row+idx, column=col_start, value=row['Instrument'])
            self.ws.cell(row=start_row+idx, column=col_start+1, value=row['Symbol'])
            self.ws.cell(row=start_row+idx, column=col_start+2, value=row['Date'])
            self.ws.cell(row=start_row+idx, column=col_start+3, value=row['Expiry'])
            self.ws.cell(row=start_row+idx, column=col_start+4, value=row['Open'])
            self.ws.cell(row=start_row+idx, column=col_start+5, value=row['High'])
            self.ws.cell(row=start_row+idx, column=col_start+6, value=row['Low'])
            self.ws.cell(row=start_row+idx, column=col_start+7, value=row['Close'])
            self.ws.cell(row=start_row+idx, column=col_start+8, value=row['LTP'])
            self.ws.cell(row=start_row+idx, column=col_start+9, value=row['Settle_Price'])
            self.ws.cell(row=start_row+idx, column=col_start+10, value=row['Contracts'])
            self.ws.cell(row=start_row+idx, column=col_start+11, value=row['Turnover_Lacs'])
            self.ws.cell(row=start_row+idx, column=col_start+12, value=row['Open_Interest'])
            self.ws.cell(row=start_row+idx, column=col_start+13, value=row['Change_OI'])
            self.ws.cell(row=start_row+idx, column=col_start+14, value=row['OI_Contracts'])
            self.ws.cell(row=start_row+idx, column=col_start+15, value=row['Change_OI_Contracts'])
    
    def apply_formatting(self, col_start):
        """Apply formatting to contract columns"""
        
        # Style header cells
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Symbol header
        self.ws.cell(row=1, column=col_start).fill = yellow_fill
        self.ws.cell(row=1, column=col_start).font = Font(bold=True)
        
        # Set column widths
        for i in range(18):
            col_letter = get_column_letter(col_start + i)
            if i in [2, 3]:  # Date and Expiry columns
                self.ws.column_dimensions[col_letter].width = 15
            elif i in [4, 5, 6, 7, 8, 9]:  # Price columns
                self.ws.column_dimensions[col_letter].width = 12
            elif i in [10, 14, 15]:  # Contract columns
                self.ws.column_dimensions[col_letter].width = 14
            else:
                self.ws.column_dimensions[col_letter].width = 11
        
        # Format date columns
        for row in range(6, self.ws.max_row + 1):
            date_cell = self.ws.cell(row=row, column=col_start+2)
            date_cell.number_format = 'DD-MMM-YY'
            
            expiry_cell = self.ws.cell(row=row, column=col_start+3)
            expiry_cell.number_format = 'DD-MMM-YY'
            
            # Format number columns
            for col_offset in [4, 5, 6, 7, 8, 9]:  # Price columns
                self.ws.cell(row=row, column=col_start+col_offset).number_format = '0.00'
            
            for col_offset in [11, 12, 13]:  # Turnover and OI value columns
                self.ws.cell(row=row, column=col_start+col_offset).number_format = '#,##0.00'
    


    def create_fudata_sheet(self):
        """Main method to create complete FuData sheet"""
        
        print("\n" + "="*70)
        print("Creating FuData Sheet for ABB Futures")
        print("="*70)
        
        self.create_workbook()
        
        all_contracts_data = []
        # for expiry in self.expiries:
        #     print(expiry)
        #     contract_df = self.generate_contract_data_v2(expiry)
        #     print(contract_df)
        # print(
        #     len(self.expiries),
        #     len(self.start_dates)
        # )
        # print(
        #     self.start_dates_dist.keys()
        # )
        for i, expiry in enumerate(self.start_dates_dist.keys()):
            # print(i, expiry)
            contract_df = self.generate_contract(expiry, i)
            # print(
            #     contract_df['Date'].iloc[-1]
            # )
            
            # break

        # for i, (expiry, start_date) in enumerate(zip(self.expiries, self.start_dates)):
        #     print(f"\nGenerating Contract {i+1}/11: {expiry.strftime('%d-%b-%Y')}")
            
        #     # Generate contract data
        #     contract_df = self.generate_contract_data(expiry, start_date, i)
        #     print(
        #         contract_df
        #     )# dataframe
            all_contracts_data.append(contract_df)
            
            # Calculate column position (each contract = 18 columns)
            col_start = (i * 18) + 1
            
            # print(f"  - Generated {len(contract_df)} trading days")
            # print(f"  - Writing to columns {col_start}-{col_start+17}")
            if len(self.start_date) > 0:
                start_date = self.start_date[-1]
                self.start_date.append(contract_df['Date'].iloc[-1])
            else:
                start_date = None
                self.start_date.append(contract_df['Date'].iloc[-1])
            # print(start_date)
            # contract_df['Date'].iloc[-1]
            expiry_date = pd.to_datetime(expiry).date()
            # Write header
            self.write_contract_header(col_start, expiry_date, start_date, contract_df)
            
            # Write data
            self.write_contract_data(col_start, contract_df)
            
            # Apply formatting
            self.apply_formatting(col_start)
        
        print("\n" + "="*70)
        print("Summary:")
        print("="*70)
        print(f"Total Contracts: {len(self.expiries)}")
        print(f"Total Columns: {len(self.expiries) * 18}")
        print(f"Total Trading Days: {sum(len(df) for df in all_contracts_data)}")
        
        return all_contracts_data
    
    def save_workbook(self, filename='ABB_FuData_Generated.xlsx'):
        """Save the workbook"""
        # filename = f"./future/{self.symbol}_FuData_Generated.xlsx"
        self.wb.save(filename)
        print(f"\n✓ Saved workbook: {filename}")
        return filename
    
    # def list_exp_start(self):
    #     from_dt = datetime(2025, 1, 1)
    #     today = datetime.today()

    #     # ensure list exists
    #     if not hasattr(self, "start_dates"):
    #         self.start_dates = []

    #     current = from_dt.replace(day=1)

    #     while current <= today:
    #         self.start_dates.append(current)

    #         # move to next month safely
    #         if current.month == 12:
    #             current = current.replace(year=current.year + 1, month=1)
    #         else:
    #             current = current.replace(month=current.month + 1)

    #     for start_date in self.start_dates:
    #         # print("cur_month:", start_date)

    #         # move to first day of next month safely
    #         next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)

    #         end_of_month = next_month - timedelta(days=1)
    #         # print("end_of_month:", end_of_month)

    #         current_date = start_date
    #         dates = []

    #         while current_date <= end_of_month:
    #             # print(current_date)
    #             # print(
    #             #     f"./NSE_Downloads/FO_Bhavcopy/FO_UDiFF_{current_date.strftime('%Y%m%d')}.csv"
    #             # )
    #             from pathlib import Path

    #             file_path = Path(f"./NSE_Downloads/FO_Bhavcopy/FO_UDiFF_{current_date.strftime('%Y%m%d')}.csv")

    #             if file_path.exists():
    #                 print(f"{file_path} File exists")
    #                 df = pd.read_csv(file_path)
    #                 # print(df)
    #                 df["XpryDt"] = pd.to_datetime(df["XpryDt"], errors="coerce")
    #                 # max_expiry = df["XpryDt"].max()
    #                 # print(max_expiry)
    #                 filtered_df = df[
    #                     (df["TckrSymb"] == f"{self.symbol}") 
    #                     &
    #                     (df["FinInstrmTp"] == "STF")
    #                 ]
    #                 # print("filtered_df", filtered_df)
    #                 max_expiry = filtered_df["XpryDt"].max()
    #                 print(start_date, max_expiry)
    #                 # if start_date not in self.start_dates:
    #                 #     self.start_dates.append(start_date)
    #                 if max_expiry not in self.expiries:
    #                     self.expiries.append(max_expiry)
    #             else:
    #                 pass
    #                 # print(f"{file_path} File does not exist")
    #             # dates.append(current_date)
    #             current_date += timedelta(days=1)
    #     print(
    #         self.start_dates,
    #         self.expiries
    #     )
    def list_exp_start(self, files):
        import os
        # folder_path = "./NSE_Downloads/FO_Bhavcopy"
        print(self.symbol)
        for file_path in files:
            # print(file_path)
            if file_path == 'None':
                continue
            df = pd.read_csv(file_path)
            # print(df[(df["TckrSymb"]== f"{self.symbol}") & (df["FinInstrmTp"] == "STF")])

            df["XpryDt"] = pd.to_datetime(df["XpryDt"], errors="coerce")
            max_expiry = df["XpryDt"].max()
            # print(max_expiry, self.symbol)
            filtered_df = df[
                (df["TckrSymb"] == f"{self.symbol}") 
                &
                (df["FinInstrmTp"] == "STF")
            ]
            if not filtered_df.empty:
                print(f"{self.symbol}", filtered_df)
            else:
                print("No records found for symbol: its index", self.symbol)
                filtered_df = df[
                    (df["TckrSymb"] == f"{self.symbol}") 
                    &
                    (df["FinInstrmTp"] == "IDF")
                ]
                # print(filtered_df)
            # print(
            #     filtered_df["TradDt"],
            #     "\n",
            #     filtered_df["XpryDt"]
            # )
            if filtered_df.empty:
                # print("no index no record", f"{self.symbol}")
                continue
            # return
            # print(
            #     filtered_df["NewBrdLotQty"].iloc[0], "$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"

            # )
            self.lot_size = filtered_df["NewBrdLotQty"].iloc[0]
            for xd in filtered_df["XpryDt"]:
                # print(
                #     (filtered_df[filtered_df["XpryDt"]==f"{xd.date()}"])["TradDt"],
                #     type((filtered_df[filtered_df["XpryDt"]==f"{xd.date()}"])["TradDt"])
                # )
                key = xd.strftime("%Y-%m-%d")
                mask = filtered_df["XpryDt"] == xd
                # print(mask)
                trad_dates = pd.to_datetime(filtered_df.loc[mask, "TradDt"]).dt.date
                # print(trad_dates.tolist())
                # pd.to_datetime(value).date()
                self.start_dates_dist[key].extend(trad_dates.tolist())
                # value = (filtered_df[filtered_df["XpryDt"]==f"{xd.date()}"])["TradDt"]
                # self.start_dates_dist[key].append(
                #     pd.to_datetime(value).date()
                # )
                # self.start_dates_dist[key].append((filtered_df[filtered_df["XpryDt"]==f"{xd.date()}"])["TradDt"])
                # print(type(xd))
                # print(filtered_df[filtered_df["XpryDt"]==f"{xd.date()}"])
                # if xd not in self.expiries:
                    # for sd in (filtered_df[filtered_df["XpryDt"]==f"{xd.date()}"])["TradDt"]:
                    #     print("start_dates", sd)
                    #     print("expiries", xd)
        #                 self.start_dates.append(sd)
                    # self.expiries.append(xd)
        # print(
        #     self.start_dates,
        #     self.expiries
        # )
        
        print(
            self.start_dates_dist
        )
            


def main():
    """Main execution"""

    # print()
    
    # # Create FuData sheet
    creator = FuDataSheetCreator(symbol='ZYDUSLIFE', lot_size=125)

    creator.list_exp_start()
    
    # Generate all contracts
    contracts_data = creator.create_fudata_sheet()
    
    # Save workbook
    output_file = creator.save_workbook('ABB_FuData_Generated.xlsx')
    
    # # Also save as CSV for easy access
    # print("\nExporting to CSV files...")
    # for i, (contract_df, expiry) in enumerate(zip(contracts_data, creator.expiries)):
    #     csv_filename = f'Contract_{expiry.strftime("%Y%m%d")}.csv'
    #     contract_df.to_csv(csv_filename, index=False)
    #     print(f"  ✓ {csv_filename}")
    
    # # Create combined CSV
    # combined_df = pd.concat(contracts_data, ignore_index=True)
    # combined_df.to_csv('All_Contracts_Combined.csv', index=False)
    # print(f"  ✓ All_Contracts_Combined.csv")
    
    # print("\n" + "="*70)
    # print("✅ FuData Sheet Creation Complete!")
    # print("="*70)
    # print(f"Main file: {output_file}")
    # print(f"Total records: {len(combined_df)}")
    # print("="*70)


if __name__ == "__main__":
    main()
