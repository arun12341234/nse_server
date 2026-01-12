#!/usr/bin/env python3
"""
EqData Sheet Generator for ABB Stock
Creates complete Equity Daily Data sheet with realistic OHLC, volume, and delivery data

Features:
- Realistic price movements with trends
- Volume patterns with spikes
- Delivery percentage calculations
- Full Excel formatting
- Technical indicators

Author: Arun - Aryan Tech World Private Limited
Date: December 15, 2025
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from collections import defaultdict

class EqDataSheetCreator:
    """Create EqData sheet with daily equity trading data"""
    
    def __init__(self, symbol='ABB', series='EQ'):
        self.symbol = symbol
        self.series = series
        
        # Trading parameters
        self.base_price = 6500
        self.volatility = 0.02  # 2% daily volatility
        self.trend_drift = 0.0005  # Small upward drift
        
        # Volume parameters
        self.avg_volume = 400000
        self.volume_volatility = 0.3
        
        # Date range
        self.start_date = datetime(2023, 12, 28)
        self.end_date = datetime(2025, 9, 12)
        
        self.wb = None
        self.ws = None

        self.start_dates_dist = []
        
    def generate_business_days(self):
        """Generate business days (excluding weekends)"""
        dates = pd.bdate_range(start=self.start_date, end=self.end_date)
        return dates
    
    def generate_price_data(self, n_days):
        """
        Generate realistic OHLC price data with trends
        
        Returns:
            DataFrame with Open, High, Low, Close, Prev_Close
        """
        
        prices = []
        prev_close = self.base_price
        
        for i in range(n_days):
            # Determine trend phase
            cycle_pos = (i % 60) / 60  # 60-day cycle
            
            if cycle_pos < 0.3:
                # Uptrend phase
                trend = 0.002
            elif cycle_pos < 0.6:
                # Sideways phase
                trend = 0.0
            else:
                # Downtrend phase
                trend = -0.0015
            
            # Random walk with trend
            daily_return = np.random.normal(trend, self.volatility)
            
            # Calculate close price
            close = prev_close * (1 + daily_return)
            
            # Generate OHLC around close
            daily_range = close * np.random.uniform(0.015, 0.03)  # 1.5-3% daily range
            
            open_price = prev_close * (1 + np.random.uniform(-0.005, 0.005))
            
            # High and Low based on the range
            if daily_return > 0:
                # Bullish day
                high = max(open_price, close) + np.random.uniform(0, daily_range * 0.5)
                low = min(open_price, close) - np.random.uniform(0, daily_range * 0.3)
            else:
                # Bearish day
                high = max(open_price, close) + np.random.uniform(0, daily_range * 0.3)
                low = min(open_price, close) - np.random.uniform(0, daily_range * 0.5)
            
            # Ensure OHLC consistency
            high = max(high, open_price, close)
            low = min(low, open_price, close)
            
            # Last price (typically close to close)
            last = close + np.random.uniform(-daily_range * 0.05, daily_range * 0.05)
            
            # Average price (VWAP approximation)
            average = (open_price + high + low + close) / 4
            
            prices.append({
                'Prev_Close': round(prev_close, 2),
                'Open': round(open_price, 2),
                'High': round(high, 2),
                'Low': round(low, 2),
                'Last': round(last, 2),
                'Close': round(close, 2),
                'Average': round(average, 2)
            })
            
            prev_close = close
        
        return pd.DataFrame(prices)
    
    def generate_volume_data(self, n_days, price_data):
        """
        Generate realistic volume and delivery data
        
        Returns:
            DataFrame with Traded_Qty, Deliverable_Qty, Delivery_Pct
        """
        
        volume_data = []
        
        for i in range(n_days):
            # Base volume with randomness
            base_vol = self.avg_volume * np.random.lognormal(0, self.volume_volatility)
            
            # Volume spikes on volatile days
            price_change = abs(price_data.iloc[i]['Close'] - price_data.iloc[i]['Prev_Close'])
            price_change_pct = price_change / price_data.iloc[i]['Prev_Close']
            
            if price_change_pct > 0.03:  # 3% move
                base_vol *= np.random.uniform(1.5, 3.0)  # Volume spike
            
            # Occasional ultra-high volume days
            if np.random.random() < 0.05:  # 5% chance
                base_vol *= np.random.uniform(3.0, 8.0)
            
            traded_qty = int(base_vol)
            traded_qty = max(traded_qty, 10000)  # Minimum volume
            
            # Delivery percentage (typically 30-60%)
            # Higher delivery on uptrends
            if price_data.iloc[i]['Close'] > price_data.iloc[i]['Prev_Close']:
                delivery_pct = np.random.uniform(40, 65)
            else:
                delivery_pct = np.random.uniform(30, 50)
            
            # Occasional very high delivery days
            if np.random.random() < 0.1:  # 10% chance
                delivery_pct = np.random.uniform(60, 75)
            
            deliverable_qty = int(traded_qty * delivery_pct / 100)
            
            # Turnover calculation
            # Turnover = Volume * Average Price / 100000 (for Lacs)
            turnover = (traded_qty * price_data.iloc[i]['Average']) / 100000
            
            volume_data.append({
                'Traded_Qty': traded_qty,
                'Deliverable_Qty': deliverable_qty,
                'Delivery_Pct': round(delivery_pct, 2),
                'Turnover_Lacs': round(turnover, 2)
            })
        
        return pd.DataFrame(volume_data)
    
    def create_equity_data(self):
        """Generate complete equity data"""
        
        print("\n" + "="*70)
        print("Generating Equity Daily Data for ABB")
        print("="*70)
        print("2")
        
        # Generate dates
        dates = self.start_dates_dist
        # print(
        #     dates,"hello"
        # )
        # self.generate_business_days()
        n_days = len(dates)
        p_data = []
        for i, date in enumerate(dates):
            print(i, date)
            # print(
            #     pd.to_datetime(date, format="%d-%b-%Y").strftime("%Y%m%d")
            # )
            from pathlib import Path
            file_path = Path(date)
            # Path(f"../NSE_Downloads/eq_security/ABB_Equity_{pd.to_datetime(date, format='%d-%b-%Y').strftime('%Y%m%d')}.csv")

            if file_path.exists():
                print(f"{file_path} File exists")
                df = pd.read_csv(file_path)
                print(f's {df["TckrSymb"]} e')
                filter_df = df[
                    (df["TckrSymb"] == self.symbol)
                    &
                    (df["SctySrs"] == 'EQ')
                ]
                print(filter_df["TckrSymb"])
                if filter_df.empty:
                    continue
                # print(
                #     type(df),
                #     type(df.to_json(orient="records", indent=2))
                # )
                import json
                price_data = json.loads(df.to_json(orient="records"))[0]
                # print(
                #     type(price_data),
                #     price_data[0]
                # )
                # volume_data = self.generate_volume_data(n_days, price_data)
                # print(
                #     price_data,
                #     volume_data
                # )
                # {
                #     "CH_SYMBOL":"ABB",
                #     "CH_SERIES":"EQ",
                #     "mTIMESTAMP":"01-Jan-2024",
                #     "CH_PREVIOUS_CLS_PRICE":4674.85,
                #     "CH_OPENING_PRICE":4674.85,
                #     "CH_TRADE_HIGH_PRICE":4715,
                #     "CH_TRADE_LOW_PRICE":4653.8,
                #     "CH_LAST_TRADED_PRICE":4668,
                #     "CH_CLOSING_PRICE":4682.2,
                #     "VWAP":4683.19,
                #     "CH_TOT_TRADED_QTY":84544,
                #     "CH_TOT_TRADED_VAL":395935611.6999999881,
                #     "CH_TOTAL_TRADES":12510,
                #     "CH_TIMESTAMP":"2023-12-31T18:30:00.000+00:00",
                #     "COP_DELIV_QTY":33919,
                #     "COP_DELIV_PERC":40.12
                # }
                # print(price_data, type(price_data))
                _list = {
                    'Symbol': price_data['CH_SYMBOL'],
                    'Series': price_data['CH_SERIES'],
                    'Date': price_data['mTIMESTAMP'],
                    'Traded_Qty': price_data['CH_TOT_TRADED_QTY'],
                    'Deliverable_Qty': price_data['COP_DELIV_QTY'],
                    'Delivery_Pct': price_data['COP_DELIV_PERC'],
                    'Prev_Close': price_data['CH_PREVIOUS_CLS_PRICE'],
                    'Open': price_data['CH_OPENING_PRICE'],
                    'High': price_data['CH_TRADE_HIGH_PRICE'],
                    'Low': price_data['CH_TRADE_LOW_PRICE'],
                    'Last': price_data['CH_LAST_TRADED_PRICE'],
                    'Close': price_data['CH_CLOSING_PRICE'],
                    'Average': price_data['CH_CLOSING_PRICE'],
                    'Total_Traded_Qty': price_data['CH_TOT_TRADED_QTY'],  # Same as Traded_Qty
                    'Turnover_Lacs': int(price_data['CH_TOT_TRADED_VAL']/100000)
                }

                print(
                    type(_list)
                )
                p_data.append(_list)
                # eq_data = pd.DataFrame()
                # print(eq_data)
        #
        eq_data = pd.DataFrame(p_data)
        eq_data['Average_Traded_Qty'] = eq_data['Traded_Qty'].expanding().mean()
        eq_data['Average_Deliverable_Qty'] = eq_data['Deliverable_Qty'].expanding().mean()
        eq_data['Average_Delivery_Pct'] = eq_data['Delivery_Pct'].expanding().mean()
        print(eq_data)
        return eq_data
        
        print(f"\nGenerating data for {n_days} trading days...")
        print(f"Date Range: {dates[0].strftime('%Y-%m-%d')} to {dates[-1].strftime('%Y-%m-%d')}")
        
        # Generate price data
        price_data = self.generate_price_data(n_days)
        print(f"✓ Generated price data (OHLC) {price_data}")
        

        # Generate volume data
        volume_data = self.generate_volume_data(n_days, price_data)
        print(f"✓ Generated volume and delivery data")
        
        # Combine all data
        eq_data = pd.DataFrame({
            'Symbol': self.symbol,
            'Series': self.series,
            'Date': dates,
            'Traded_Qty': volume_data['Traded_Qty'],
            'Deliverable_Qty': volume_data['Deliverable_Qty'],
            'Delivery_Pct': volume_data['Delivery_Pct'],
            'Prev_Close': price_data['Prev_Close'],
            'Open': price_data['Open'],
            'High': price_data['High'],
            'Low': price_data['Low'],
            'Last': price_data['Last'],
            'Close': price_data['Close'],
            'Average': price_data['Average'],
            'Total_Traded_Qty': volume_data['Traded_Qty'],  # Same as Traded_Qty
            'Turnover_Lacs': volume_data['Turnover_Lacs']
        })
        
        # Calculate additional metrics
        eq_data['Average_Traded_Qty'] = eq_data['Traded_Qty'].expanding().mean()
        eq_data['Average_Deliverable_Qty'] = eq_data['Deliverable_Qty'].expanding().mean()
        eq_data['Average_Delivery_Pct'] = eq_data['Delivery_Pct'].expanding().mean()
        
        # Five Day Price Down flag
        eq_data['Five_Day_Down'] = 'No'
        for i in range(4, len(eq_data)):
            last_5 = eq_data['Close'].iloc[i-4:i+1]
            if all(last_5.diff().dropna() < 0):
                eq_data.loc[i, 'Five_Day_Down'] = 'Yes'
        
        print(f"✓ Calculated additional metrics")
        
        return eq_data
    
    def create_workbook(self, eq_data, filename):
        """Create Excel workbook with formatting"""
        
        print("\nCreating Excel workbook...")
        from openpyxl import load_workbook
        from openpyxl import Workbook
        import os
        if os.path.exists(filename):
            # Load existing workbook (keeps abc)
            self.wb = load_workbook(filename)
            print("exist")
        else:
            # Create new workbook
            self.wb = Workbook()
        # self.wb = Workbook()
        # self.ws = self.wb.active
        # self.ws.title = "EqData11"
        self.ws = self.wb.create_sheet(title="EqData")
        
        # Row 1 - Title
        title_cell = self.ws.cell(row=1, column=7, value="Equity Daily Data")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge title cells
        self.ws.merge_cells('G1:I1')
        
        # Row 2 - Column Headers
        headers = [
            'Symbol', 'Series', 'Date', 'Traded Quantity', 'Deliverable Qty',
            '% Dly Qty to Traded Qty ', ' Prev Close', 'Open Price', 'High Price',
            'Low Price', 'Last Price', 'Close Price', 'Average Price',
            'Total Traded Quantity', 'Turnover in Lacs', 'Average Traded Qty',
            '', 'Five Day Price Down Continuously', 'Average of Deliverable Qty',
            'Average of % Dly Qty to Traded Qty '
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Data rows (starting from row 3)
        for row_idx, (_, row_data) in enumerate(eq_data.iterrows(), start=3):
            self.ws.cell(row=row_idx, column=1, value=row_data['Symbol'])
            self.ws.cell(row=row_idx, column=2, value=row_data['Series'])
            self.ws.cell(row=row_idx, column=3, value=row_data['Date'])
            self.ws.cell(row=row_idx, column=4, value=row_data['Traded_Qty'])
            self.ws.cell(row=row_idx, column=5, value=row_data['Deliverable_Qty'])
            self.ws.cell(row=row_idx, column=6, value=row_data['Delivery_Pct'])
            self.ws.cell(row=row_idx, column=7, value=row_data['Prev_Close'])
            self.ws.cell(row=row_idx, column=8, value=row_data['Open'])
            self.ws.cell(row=row_idx, column=9, value=row_data['High'])
            self.ws.cell(row=row_idx, column=10, value=row_data['Low'])
            self.ws.cell(row=row_idx, column=11, value=row_data['Last'])
            self.ws.cell(row=row_idx, column=12, value=row_data['Close'])
            self.ws.cell(row=row_idx, column=13, value=row_data['Average'])
            self.ws.cell(row=row_idx, column=14, value=row_data['Total_Traded_Qty'])
            self.ws.cell(row=row_idx, column=15, value=row_data['Turnover_Lacs'])
            self.ws.cell(row=row_idx, column=16, value=row_data['Average_Traded_Qty'])
            self.ws.cell(row=row_idx, column=17, value='')
            # self.ws.cell(row=row_idx, column=18, value=row_data['Five_Day_Down'])
            # self.ws.cell(row=row_idx, column=19, value=row_data['Average_Deliverable_Qty'])
            # self.ws.cell(row=row_idx, column=20, value=row_data['Average_Delivery_Pct'])
        
        # Apply formatting
        self.apply_formatting(len(eq_data))
        
        print("✓ Excel workbook created with formatting")
    
    def apply_formatting(self, n_rows):
        """Apply formatting to the sheet"""
        
        # Set column widths
        column_widths = {
            'A': 8, 'B': 8, 'C': 12, 'D': 14, 'E': 14,
            'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12,
            'K': 12, 'L': 12, 'M': 12, 'N': 14, 'O': 15,
            'P': 16, 'Q': 5, 'R': 15, 'S': 18, 'T': 22
        }
        
        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width
        
        # Format dates (column C)
        for row in range(3, n_rows + 3):
            date_cell = self.ws.cell(row=row, column=3)
            date_cell.number_format = 'DD-MMM-YY'
        
        # Format numbers
        for row in range(3, n_rows + 3):
            # Quantities - no decimals
            for col in [4, 5, 14]:
                self.ws.cell(row=row, column=col).number_format = '#,##0'
            
            # Percentages - 2 decimals
            self.ws.cell(row=row, column=6).number_format = '0.00'
            
            # Prices - 2 decimals
            for col in [7, 8, 9, 10, 11, 12, 13]:
                self.ws.cell(row=row, column=col).number_format = '0.00'
            
            # Turnover - 2 decimals
            self.ws.cell(row=row, column=15).number_format = '#,##0.00'
            
            # Averages - 2 decimals
            for col in [16, 19, 20]:
                self.ws.cell(row=row, column=col).number_format = '#,##0.00'
        
        # Conditional formatting for Five Day Down
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in range(3, n_rows + 3):
            cell = self.ws.cell(row=row, column=18)
            if cell.value == 'Yes':
                cell.fill = red_fill
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(2, n_rows + 3):
            for col in range(1, 21):
                self.ws.cell(row=row, column=col).border = thin_border
        
        print("✓ Applied formatting (dates, numbers, borders)")
    
    def save_workbook(self, filename='ABB_EqData_Generated.xlsx'):
        """Save the workbook"""
        # from openpyxl import load_workbook
        # from openpyxl import Workbook
        # import os
        # if os.path.exists(filename):
        #     # Load existing workbook (keeps abc)
        #     self.wb = load_workbook(filename)
        # else:
        #     # Create new workbook
            # self.wb = Workbook()
        self.wb.save(filename)
        print(f"\n✓ Saved workbook: {filename}")
        return filename
    
    def create_eqdata_sheet(self, filename):
        """Main method to create complete EqData sheet"""
        
        # Generate equity data
        eq_data = self.create_equity_data()
        print("1")
        
        # Create Excel workbook
        self.create_workbook(eq_data, filename)
        
        # # Print summary
        # print("\n" + "="*70)
        # print("Summary:")
        # print("="*70)
        # print(f"Symbol: {self.symbol}")
        # print(f"Trading Days: {len(eq_data)}")
        # print(f"Date Range: {eq_data['Date'].min().strftime('%Y-%m-%d')} to {eq_data['Date'].max().strftime('%Y-%m-%d')}")
        # print(f"Opening Price: ₹{eq_data['Close'].iloc[0]:.2f}")
        # print(f"Closing Price: ₹{eq_data['Close'].iloc[-1]:.2f}")
        # print(f"Change: ₹{eq_data['Close'].iloc[-1] - eq_data['Close'].iloc[0]:.2f}")
        # print(f"Change %: {((eq_data['Close'].iloc[-1] / eq_data['Close'].iloc[0]) - 1) * 100:.2f}%")
        # print(f"High: ₹{eq_data['High'].max():.2f}")
        # print(f"Low: ₹{eq_data['Low'].min():.2f}")
        # print(f"Total Volume: {eq_data['Traded_Qty'].sum():,} shares")
        # print(f"Avg Volume: {eq_data['Traded_Qty'].mean():,.0f} shares")
        # print(f"Avg Delivery %: {eq_data['Delivery_Pct'].mean():.2f}%")
        
        # return eq_data

    def list_exp_start(self, dates):
        self.start_dates_dist = dates
        # import os
        # # folder_path = "../NSE_Downloads/eq_security"
        # # files = os.listdir(folder_path)
        # # print(files)
        # for file_path in files:
        #     if file_path == 'None':
        #         continue
        #     else:
        #         print(file_path)
        #         df = pd.read_csv(file_path)
        #         print(df)

        #     try:
        #         df = pd.read_csv(f"{folder_path}/"+file_path)
        #         print(df)
        #         print(
        #             df["mTIMESTAMP"]
        #         )
        #         filtered_df = df["mTIMESTAMP"]
        #         for ts in filtered_df:
        #             print(ts)
        #             self.start_dates_dist.append(ts)
        #     except Exception as e:
        #         print(e)
        # print(
        #     self.start_dates_dist
        # )



def main():
    """Main execution"""
    
    # Create EqData sheet
    creator = EqDataSheetCreator(symbol='ABB', series='EQ')
    
    # Customize if needed
    # creator.base_price = 6500
    # creator.start_date = datetime(2024, 1, 1)
    # creator.end_date = datetime(2025, 12, 31)
    
    # Generate equity data
    eq_data = creator.create_eqdata_sheet()
    
    # Save workbook
    output_file = creator.save_workbook('ABB_EqData_Generated.xlsx')
    
    # Also save as CSV
    # print("\nExporting to CSV...")
    # csv_file = 'ABB_EqData_Generated.csv'
    # eq_data.to_csv(csv_file, index=False)
    # print(f"✓ CSV exported: {csv_file}")
    
    # print("\n" + "="*70)
    # print("✅ EqData Sheet Creation Complete!")
    # print("="*70)
    # print(f"Excel file: {output_file}")
    # print(f"CSV file: {csv_file}")
    # print(f"Total records: {len(eq_data)}")
    # print("="*70)


if __name__ == "__main__":
    main()
