import os
from flask import Flask, jsonify, request
import openpyxl  # You'll need to run: pip install openpyxl
from datetime import date, timedelta
app = Flask(__name__)

# The base directory where folders will be born
BASE_DIR = "Nse_files"

@app.route('/api/<year>', methods=['GET', 'POST'])
def create_year_folder(year):
    """
    Creates a folder for the specified year.
    Usage: GET or POST /api/2024
    """
    
    # 1. Sanity Check: Ensure the year looks like a year (digits only)
    # We don't want folders named 'Robert' or 'drop_tables'.
    if not year.isdigit():
        return jsonify({
            "status": "error",
            "message": "That doesn't look like a year. Please use numbers, not emotional baggage."
        }), 400

    # 2. Construct the full path
    # os.path.join is our friend for cross-platform compatibility
    folder_path = os.path.join(BASE_DIR, year)

    try:
        # 3. Check if it exists, then create
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
            return jsonify({
                "status": "success",
                "message": f"Folder '{year}' created! It's empty and waiting for memories."
            }), 201
        else:
            return jsonify({
                "status": "ignored",
                "message": f"Folder '{year}' already exists. We didn't want to overwrite history."
            }), 200

    except Exception as e:
        # If the server catches on fire, tell the user why
        return jsonify({
            "status": "error",
            "message": f"Something went wrong: {str(e)}"
        }), 500

@app.route('/api/<year>/track_date', methods=['GET', 'POST'])
def create_track_date(year):
    """
    Creates a tracking Excel sheet in the year folder.
    Usage: GET or POST /api/2024/track_date
    """
    
    # 1. Sanity Check
    if not year.isdigit():
        return jsonify({
            "status": "error",
            "message": "That doesn't look like a year. Please use numbers."
        }), 400

    folder_path = os.path.join(BASE_DIR, year)
    file_path = os.path.join(folder_path, 'tracking.xlsx')

    try:
        # Ensure the year folder exists first
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # 2. Check if Excel exists, then create
        if not os.path.exists(file_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Create headers: date, file_1, file_2 ... file_11
            headers = ['date'] + [f'file_{i}' for i in range(1, 12)]
            ws.append(headers)
            
            wb.save(file_path)
            
            return jsonify({
                "status": "success",
                "message": f"Tracking sheet created at '{file_path}'. It's ready to judge your files."
            }), 201
        else:
            return jsonify({
                "status": "ignored",
                "message": "Tracking sheet already exists. We backed away slowly."
            }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Excel creation failed: {str(e)}"
        }), 500

@app.route('/api/<year>/tracking', methods=['GET'])
def read_tracking_sheet(year):
    """
    Reads the content of the tracking Excel sheet.
    Usage: GET /api/2024/tracking
    """
    
    # 1. Sanity Check
    if not year.isdigit():
        return jsonify({
            "status": "error",
            "message": "That year is suspicious. Try digits."
        }), 400

    file_path = os.path.join(BASE_DIR, year, 'tracking.xlsx')

    # 2. Check if the file is actually there
    if not os.path.exists(file_path):
        return jsonify({
            "status": "error",
            "message": "The Oracle (Excel file) could not be found. Did you create it first?"
        }), 404

    try:
        # 3. Read the file
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        data = []
        # Get headers from the first row
        headers = [cell.value for cell in ws[1]]
        
        # Iterate through the rest of the rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Zip headers with values to make a nice dictionary
            row_data = dict(zip(headers, row))
            data.append(row_data)
            
        return jsonify({
            "status": "success",
            "data": data,
            "message": "Here is the data you requested, served fresh from the spreadsheet."
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Reading failed. The file refused to talk: {str(e)}"
        }), 500

@app.route('/api/<year>/tracking/<date>', methods=['GET', 'POST'])
def manage_tracking_by_date(year, date):
    """
    GET: Finds a specific row by date.
    POST: Updates the specific row for that date with JSON data OR creates it if missing.
    Usage: GET /api/2024/tracking/2024-01-01
           POST /api/2024/tracking/2024-01-01 with JSON body {"file_1": "Found it"}
    """
    
    # 1. Sanity Check
    if not year.isdigit():
        return jsonify({
            "status": "error",
            "message": "Year invalid. Time is a construct, but digits are required."
        }), 400

    file_path = os.path.join(BASE_DIR, year, 'tracking.xlsx')

    # 2. Check if the file exists
    if not os.path.exists(file_path):
        return jsonify({
            "status": "error",
            "message": "No tracking file found. It's like looking for a needle in a vacuum."
        }), 404

    try:
        # 3. Load the workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        target_row_idx = None
        found_data = None
        
        # We search for the row index (for POST) and data (for GET)
        # We start at row 2 because row 1 is headers
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = dict(zip(headers, row))
            
            # Convert to string to compare safely
            row_date_str = str(row_data.get('date', ''))
            
            if row_date_str == date:
                target_row_idx = idx
                found_data = row_data
                break
        
        # --- POST Logic (Update or Append) ---
        if request.method == 'POST':
            updates = request.get_json()
            if not updates:
                return jsonify({
                    "status": "error",
                    "message": "You sent a POST request but no data. I'm not a mind reader."
                }), 400

            # Map headers to column indices (1-based for openpyxl)
            header_map = {name: i + 1 for i, name in enumerate(headers)}
            
            # Scenario A: Update existing row
            if target_row_idx:
                changes_made = 0
                for key, value in updates.items():
                    if key in header_map:
                        col_idx = header_map[key]
                        ws.cell(row=target_row_idx, column=col_idx, value=value)
                        changes_made += 1
                
                wb.save(file_path)
                return jsonify({
                    "status": "success",
                    "message": f"Updated {changes_made} fields for {date}. History has been rewritten."
                }), 200
            
            # Scenario B: Append new row
            else:
                next_row = ws.max_row + 1
                
                # Automatically set the 'date' column based on the URL parameter
                if 'date' in header_map:
                    ws.cell(row=next_row, column=header_map['date'], value=date)
                
                # Fill in the rest from the JSON body
                for key, value in updates.items():
                    if key in header_map:
                        col_idx = header_map[key]
                        ws.cell(row=next_row, column=col_idx, value=value)
                
                wb.save(file_path)
                return jsonify({
                    "status": "success",
                    "message": f"New entry created for {date}. Welcome to the timeline."
                }), 201

        # --- GET Logic (Read) ---
        else:
            if found_data:
                return jsonify({
                    "status": "success",
                    "data": found_data,
                    "message": f"Eureka! We found the entry for {date}."
                }), 200
            else:
                return jsonify({
                    "status": "error",
                    "message": f"The date {date} is missing from the chronicles."
                }), 404

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Operation failed. The spreadsheet is being stubborn: {str(e)}"
        }), 500

@app.route('/api/download/<int:year>', methods=['GET'])
def get_year_dates(year):
    """
    Generates dates for the requested year + previous 3 months.
    - If requested year == current year: Ends at TODAY.
    - Otherwise: Ends at December 31st of the requested year.
    """
    try:
        # Get today's date
        today = date.today()

        # 1. Start Date: Always October 1st of the previous year
        start_date = date(year - 1, 10, 1)

        # 2. End Date Logic
        if year == today.year:
            # If the user asks for the current running year, stop at today
            end_date = today - timedelta(days=1)
        else:
            # Otherwise (past or future years), go to the end of that year
            end_date = date(year, 12, 31)

        # 3. Generate the list
        all_dates = []
        current_date = start_date
        
        while current_date <= end_date:
            all_dates.append(current_date.strftime("%Y-%m-%d"))
            current_date += timedelta(days=1)

        # 4. Return Response
        return jsonify({
            "status": "success",
            "requested_year": year,
            "is_current_year": (year == today.year),
            "range_start": start_date.strftime("%Y-%m-%d"),
            "range_end": end_date.strftime("%Y-%m-%d"),
            "total_days": len(all_dates),
            "dates": all_dates
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 400

if __name__ == '__main__':
    # Create the base directory if it doesn't exist
    if not os.path.exists(BASE_DIR):
        os.makedirs(BASE_DIR)
        print(f"Created base directory: {BASE_DIR}")

    # Run the app
    print("Server running! Send requests to http://127.0.0.1:5000/api/<year>")
    app.run(debug=True, port=5000)