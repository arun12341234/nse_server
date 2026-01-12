#!/usr/bin/env python3
"""
EQUITY Data Multi-Sheet Filler - One Sheet Per Symbol
======================================================
Creates separate Excel sheets for each stock symbol
All sheets follow the exact EQUITY_Data.xlsx template format
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')


class EquityMultiSheetFiller:
    """Create Excel with separate sheet for each stock symbol"""
    
    def __init__(self, file_list, template_path):
        """Initialize with file list and template"""
        self.file_list = [f for f in file_list if f != 'None']
        self.template_path = template_path
        self.all_data = []
        self.symbols_data = {}
        
    def load_all_files(self):
        """Load all CSV files"""
        print("=" * 80)
        print("📂 LOADING NSE FILES")
        print("=" * 80)
        
        loaded_count = 0
        
        for file_path in self.file_list:
            try:
                # Extract date from filename
                filename = os.path.basename(file_path)
                date_str = filename.replace('sec_bhavdata_full_', '').replace('.csv', '')
                
                # Read CSV
                df = pd.read_csv(file_path)
                
                # Clean column names and data
                df.columns = df.columns.str.strip()
                for col in df.select_dtypes(include=['object']).columns:
                    df[col] = df[col].str.strip()
                
                # Add trade date
                df['TRADE_DATE'] = pd.to_datetime(date_str, format='%Y%m%d')
                
                # Convert numeric columns
                numeric_cols = ['PREV_CLOSE', 'OPEN_PRICE', 'HIGH_PRICE', 'LOW_PRICE', 
                               'LAST_PRICE', 'CLOSE_PRICE', 'AVG_PRICE', 'TTL_TRD_QNTY',
                               'TURNOVER_LACS', 'NO_OF_TRADES', 'DELIV_QTY', 'DELIV_PER']
                
                for col in numeric_cols:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce')
                
                # Filter for EQ series only
                df_eq = df[df['SERIES'] == 'EQ'].copy()
                
                self.all_data.append(df_eq)
                loaded_count += 1
                
                if loaded_count % 10 == 0:
                    print(f"  ✓ Loaded {loaded_count} files...")
                    
            except Exception as e:
                print(f"  ⚠ Error loading {file_path}: {str(e)}")
                continue
        
        print()
        print(f"✓ Successfully loaded {loaded_count} files")
        print(f"✓ Total records: {sum(len(df) for df in self.all_data):,}")
        print()
        
        return self.all_data
    
    def organize_by_symbol(self):
        """Organize data by stock symbol"""
        print("=" * 80)
        print("📊 ORGANIZING DATA BY SYMBOL")
        print("=" * 80)
        
        # Combine all data
        combined_df = pd.concat(self.all_data, ignore_index=True)
        
        # Sort by symbol and date
        combined_df = combined_df.sort_values(['SYMBOL', 'TRADE_DATE']).reset_index(drop=True)
        
        # Group by symbol
        self.symbols_data = {}
        for symbol in combined_df['SYMBOL'].unique():
            self.symbols_data[symbol] = combined_df[combined_df['SYMBOL'] == symbol].copy()
        
        print(f"✓ Found {len(self.symbols_data)} unique stock symbols")
        print()
        
        # Show top 10 symbols by record count
        symbol_counts = {sym: len(df) for sym, df in self.symbols_data.items()}
        top_symbols = sorted(symbol_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        
        print("Top 10 symbols by number of records:")
        for symbol, count in top_symbols:
            print(f"  {symbol:<15} {count:>5} records")
        print()
        
        return self.symbols_data
    
    def create_multi_sheet_excel(self, output_path, max_symbols=None):
        """Create Excel with one sheet per symbol"""
        print("=" * 80)
        print("📝 CREATING MULTI-SHEET EXCEL - ONE SHEET PER SYMBOL")
        print("=" * 80)
        
        # Load template
        wb_template = openpyxl.load_workbook(self.template_path)
        ws_template = wb_template.active
        
        # Create new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Limit symbols if requested
        symbols_to_process = list(self.symbols_data.keys())
        if max_symbols:
            symbols_to_process = symbols_to_process[:max_symbols]
            print(f"  ℹ Processing first {max_symbols} symbols only")
        
        print(f"  ℹ Creating sheets for {len(symbols_to_process)} symbols")
        print()
        
        # Create sheet for each symbol
        for idx, symbol in enumerate(symbols_to_process, 1):
            df_symbol = self.symbols_data[symbol]
            
            # Create sheet name (Excel max 31 chars)
            sheet_name = symbol[:31]
            
            # Create new sheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Fill the sheet with data
            self._fill_sheet_for_symbol(ws, ws_template, df_symbol, symbol)
            
            # Progress
            if idx % 10 == 0:
                print(f"  ✓ Created {idx} sheets...")
        
        # Save workbook
        wb.save(output_path)
        
        print()
        print(f"✓ Excel file created: {output_path}")
        print(f"✓ Total sheets: {len(symbols_to_process)}")
        print(f"✓ Each sheet contains data for one stock symbol")
        print()
        
        return output_path
    
    def _fill_sheet_for_symbol(self, ws, ws_template, df_symbol, symbol):
        """Fill a single sheet for one symbol"""
        
        # --- COPY TEMPLATE STRUCTURE (Rows 1-10) ---
        self._copy_template_structure(ws, ws_template)
        
        # --- CALCULATE AND FILL AVERAGES ---
        avg_traded_qty = df_symbol['TTL_TRD_QNTY'].mean()
        avg_deliv_qty = df_symbol['DELIV_QTY'].mean()
        avg_deliv_pct = df_symbol['DELIV_PER'].mean()
        avg_turnover = df_symbol['TURNOVER_LACS'].mean()
        
        # Fill averages (in non-merged cells)
        ws.cell(row=4, column=8).value = avg_traded_qty
        ws.cell(row=4, column=8).number_format = '#,##0.00'
        
        ws.cell(row=5, column=8).value = avg_deliv_qty
        ws.cell(row=5, column=8).number_format = '#,##0.00'
        
        ws.cell(row=6, column=8).value = avg_deliv_pct
        ws.cell(row=6, column=8).number_format = '0.00'
        
        ws.cell(row=7, column=8).value = avg_turnover
        ws.cell(row=7, column=8).number_format = '#,##0.00'
        
        # --- ROW 9: Stock identifier ---
        ws.cell(row=9, column=10).value = symbol
        
        # Date range in Q9
        date_range = f"{df_symbol['TRADE_DATE'].min().strftime('%d-%b-%Y')} to {df_symbol['TRADE_DATE'].max().strftime('%d-%b-%Y')}"
        ws.cell(row=9, column=17).value = date_range
        
        # --- DATA ROWS (Starting from Row 11) ---
        self._fill_data_rows(ws, ws_template, df_symbol, start_row=11)
        
        # --- COPY COLUMN WIDTHS ---
        for col_idx in range(1, 20):
            col_letter = get_column_letter(col_idx)
            if col_letter in ws_template.column_dimensions:
                ws.column_dimensions[col_letter].width = ws_template.column_dimensions[col_letter].width
    
    def _copy_template_structure(self, ws_dest, ws_source):
        """Copy template structure (rows 1-10) with all formatting"""
        
        # Copy rows 1-10 with all formatting (before merging)
        for row_idx in range(1, 11):
            ws_dest.row_dimensions[row_idx].height = ws_source.row_dimensions[row_idx].height
            
            for col_idx in range(1, 20):
                source_cell = ws_source.cell(row=row_idx, column=col_idx)
                dest_cell = ws_dest.cell(row=row_idx, column=col_idx)
                
                # Copy value only for non-merged cells
                if not isinstance(source_cell, openpyxl.cell.cell.MergedCell):
                    dest_cell.value = source_cell.value
                
                # Copy formatting
                if source_cell.font:
                    dest_cell.font = copy(source_cell.font)
                if source_cell.fill:
                    dest_cell.fill = copy(source_cell.fill)
                if source_cell.alignment:
                    dest_cell.alignment = copy(source_cell.alignment)
                if source_cell.border:
                    dest_cell.border = copy(source_cell.border)
                if source_cell.number_format:
                    dest_cell.number_format = source_cell.number_format
        
        # Copy merged cells
        for merged_range in ws_source.merged_cells.ranges:
            ws_dest.merge_cells(str(merged_range))
    
    def _fill_data_rows(self, ws, ws_template, df, start_row=11):
        """Fill data rows starting from specified row"""
        
        # Store template row formatting from row 11
        template_formats = {}
        for col_idx in range(3, 17):  # Columns C to P
            source_cell = ws_template.cell(row=11, column=col_idx)
            template_formats[col_idx] = {
                'font': copy(source_cell.font) if source_cell.font else None,
                'fill': copy(source_cell.fill) if source_cell.fill else None,
                'alignment': copy(source_cell.alignment) if source_cell.alignment else None,
                'border': copy(source_cell.border) if source_cell.border else None,
                'number_format': source_cell.number_format if source_cell.number_format else None,
            }
        
        # Fill data
        for idx, row_data in df.iterrows():
            row_num = start_row + idx
            
            # Column C: Date
            ws.cell(row=row_num, column=3).value = row_data['TRADE_DATE']
            
            # Column D: PREV_CLOSE
            ws.cell(row=row_num, column=4).value = row_data['PREV_CLOSE']
            
            # Column E: OPEN_PRICE
            ws.cell(row=row_num, column=5).value = row_data['OPEN_PRICE']
            
            # Column F: HIGH_PRICE
            ws.cell(row=row_num, column=6).value = row_data['HIGH_PRICE']
            
            # Column G: LOW_PRICE
            ws.cell(row=row_num, column=7).value = row_data['LOW_PRICE']
            
            # Column H: LAST_PRICE
            ws.cell(row=row_num, column=8).value = row_data['LAST_PRICE']
            
            # Column I: CLOSE_PRICE
            ws.cell(row=row_num, column=9).value = row_data['CLOSE_PRICE']
            
            # Column J: AVG_PRICE
            ws.cell(row=row_num, column=10).value = row_data['AVG_PRICE']
            
            # Column K: TTL_TRD_QNTY
            ws.cell(row=row_num, column=11).value = row_data['TTL_TRD_QNTY']
            
            # Column L: TURNOVER_LACS
            ws.cell(row=row_num, column=12).value = row_data['TURNOVER_LACS']
            
            # Column M: NO_OF_TRADES
            ws.cell(row=row_num, column=13).value = row_data['NO_OF_TRADES']
            
            # Column N: DELIVERY QTY (formula)
            ws.cell(row=row_num, column=14).value = f'=G{row_num}/J{row_num}'
            
            # Column O: DELIVERY % (formula)
            ws.cell(row=row_num, column=15).value = f'=H{row_num}/I{row_num}'
            
            # Column P: UP/DOWN (calculated value)
            if pd.notna(row_data['CLOSE_PRICE']) and pd.notna(row_data['PREV_CLOSE']) and row_data['PREV_CLOSE'] != 0:
                up_down = (row_data['CLOSE_PRICE'] - row_data['PREV_CLOSE']) / row_data['PREV_CLOSE']
                ws.cell(row=row_num, column=16).value = up_down
            
            # Apply template formatting to all cells
            for col_idx in range(3, 17):
                cell = ws.cell(row=row_num, column=col_idx)
                fmt = template_formats.get(col_idx, {})
                
                if fmt.get('font'):
                    cell.font = copy(fmt['font'])
                if fmt.get('fill'):
                    cell.fill = copy(fmt['fill'])
                if fmt.get('alignment'):
                    cell.alignment = copy(fmt['alignment'])
                if fmt.get('border'):
                    cell.border = copy(fmt['border'])
                if fmt.get('number_format'):
                    cell.number_format = fmt['number_format']
    
    def create_summary_sheet(self, wb):
        """Create a summary sheet with all symbols"""
        ws = wb.create_sheet(title='Summary', index=0)
        
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = '📊 STOCK SYMBOLS SUMMARY'
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Headers
        headers = ['Symbol', 'Records', 'Start Date', 'End Date', 'Avg Turnover (Lacs)']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row_num = 4
        for symbol, df in sorted(self.symbols_data.items()):
            ws.cell(row=row_num, column=1).value = symbol
            ws.cell(row=row_num, column=2).value = len(df)
            ws.cell(row=row_num, column=3).value = df['TRADE_DATE'].min()
            ws.cell(row=row_num, column=4).value = df['TRADE_DATE'].max()
            ws.cell(row=row_num, column=5).value = df['TURNOVER_LACS'].mean()
            ws.cell(row=row_num, column=5).number_format = '#,##0.00'
            
            row_num += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
    
    def print_summary(self):
        """Print summary statistics"""
        print("=" * 80)
        print("📊 DATA SUMMARY")
        print("=" * 80)
        
        total_records = sum(len(df) for df in self.symbols_data.values())
        
        print(f"Total Symbols:        {len(self.symbols_data):,}")
        print(f"Total Records:        {total_records:,}")
        print()
        
        # Find symbol with most/least records
        symbol_counts = {sym: len(df) for sym, df in self.symbols_data.items()}
        max_symbol = max(symbol_counts.items(), key=lambda x: x[1])
        min_symbol = min(symbol_counts.items(), key=lambda x: x[1])
        
        print(f"Most data:            {max_symbol[0]} ({max_symbol[1]} records)")
        print(f"Least data:           {min_symbol[0]} ({min_symbol[1]} records)")
        print()
        
        # Overall date range
        all_dates = pd.concat([df['TRADE_DATE'] for df in self.symbols_data.values()])
        print(f"Date Range:           {all_dates.min().strftime('%d-%b-%Y')} to {all_dates.max().strftime('%d-%b-%Y')}")
        print()
        print("=" * 80)
        print()

import requests
def main():
    """Main execution"""
    YEAR_TO_FETCH = 2026
    url = f"http://127.0.0.1:5000/api/{YEAR_TO_FETCH}/tracking"
    payload = {
            "file_8": f"data_value"
    }
    response = requests.get(url, json=payload)
    data = response.json().get("data")
    file_1_list = [item["file_8"] for item in data if "file_8" in item]
    # File list
    file_list = file_1_list
    
    print("=" * 80)
    print("🎯 EQUITY DATA - ONE SHEET PER SYMBOL")
    print("=" * 80)
    print()
    print("This script creates a SINGLE Excel file with:")
    print("  ✓ One sheet for EACH stock symbol")
    print("  ✓ Each sheet follows exact template format")
    print("  ✓ All formatting, colors, borders preserved")
    print("  ✓ Summary sheet with list of all symbols")
    print()
    print("Features:")
    print("  • Separate sheet for RELIANCE, TCS, INFY, etc.")
    print("  • Each sheet has all data for that stock")
    print("  • Automatic averages calculation per stock")
    print("  • Easy navigation between stocks")
    print()
    print("=" * 80)
    print()
    
    # When files are available, uncomment:
    # Initialize
    filler = EquityMultiSheetFiller(
        file_list=file_list,
        template_path='EQUITY_Data.xlsx'
    )
    
    # Load files
    filler.load_all_files()
    
    # Organize by symbol
    filler.organize_by_symbol()
    
    # Print summary
    filler.print_summary()
    
    # Create multi-sheet Excel
    output_file = 'EQUITY_Data_By_Symbol.xlsx'
    filler.create_multi_sheet_excel(
        output_path=output_file,
        max_symbols=None  # None for all symbols, or 100 to limit
    )
    
    # Add summary sheet
    wb = openpyxl.load_workbook(output_file)
    filler.create_summary_sheet(wb)
    wb.save(output_file)
    
    print()
    print("=" * 80)
    print("🎉 COMPLETE!")
    print("=" * 80)
    print(f"✓ File: {output_file}")
    print("✓ One sheet per stock symbol")
    print("✓ Summary sheet with all symbols")
    print("✓ Exact template format preserved")
    print()


if __name__ == "__main__":
    main()